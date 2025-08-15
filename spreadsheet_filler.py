#!/usr/bin/env python3
"""
Spreadsheet Filler v1.0
- Supports CSV / Excel / Google Sheets input
- Maps columns to a target Excel template and writes output
- Data cleaning, validation, simple formatting
- Backup, dry-run, logging
"""

import os
import sys
import argparse
import json
import shutil
import logging
from datetime import datetime
from pathlib import Path
import re

import pandas as pd
from dateutil import parser as dateparser
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

# (Optional) gspread imports if Google Sheets support is enabled
try:
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
    GS_SUPPORTED = True
except Exception:
    GS_SUPPORTED = False

# -------------------------
# Logging config
# -------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s — %(levelname)s — %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)

# -------------------------
# Utilities
# -------------------------
def backup_file(path: str):
    """Create a timestamped backup of a file (if it exists)."""
    p = Path(path)
    if not p.exists():
        return None
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = p.with_name(f"{p.stem}_backup_{stamp}{p.suffix}")
    shutil.copy2(p, backup)
    logging.info(f"Backup created: {backup}")
    return backup

def read_config(path: str):
    """Read JSON config mapping file."""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

# -------------------------
# Input loaders
# -------------------------
def load_source_file(path: str) -> pd.DataFrame:
    """Load CSV or Excel into a pandas DataFrame."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Source file not found: {path}")
    if p.suffix.lower() in [".csv"]:
        df = pd.read_csv(path)
    elif p.suffix.lower() in [".xlsx", ".xls"]:
        df = pd.read_excel(path)
    else:
        raise ValueError("Unsupported file type. Use .csv or .xlsx/.xls")
    logging.info(f"Loaded {len(df)} rows from {path}")
    return df

def load_google_sheet(spreadsheet_key: str, worksheet_name: str, creds_json: str) -> pd.DataFrame:
    """Load Google Sheet into DataFrame using service account credentials JSON path."""
    if not GS_SUPPORTED:
        raise RuntimeError("gspread / oauth2client not installed or available.")
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(creds_json, scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_key(spreadsheet_key).worksheet(worksheet_name)
    data = sheet.get_all_records()
    df = pd.DataFrame(data)
    logging.info(f"Loaded {len(df)} rows from Google Sheet {spreadsheet_key}/{worksheet_name}")
    return df

# -------------------------
# Cleaning / Validation
# -------------------------
def clean_cell(value):
    """Basic cleaning: strip strings, convert NaNs to empty, keep types."""
    if pd.isna(value):
        return None
    if isinstance(value, str):
        v = value.strip()
        # normalize multiple spaces
        v = re.sub(r"\s+", " ", v)
        return v
    return value

def parse_date(value):
    """Try to parse a date-like value into datetime.date"""
    if value is None:
        return None
    if isinstance(value, (datetime, pd.Timestamp)):
        return value
    try:
        return dateparser.parse(str(value))
    except Exception:
        return None

def run_cleaning(df: pd.DataFrame, rules: dict) -> pd.DataFrame:
    """Apply simple cleaning rules to DataFrame.
    rules may contain keys: lowercase_fields, titlecase_fields, date_fields, strip_fields
    """
    df = df.copy()
    for col in df.columns:
        df[col] = df[col].apply(clean_cell)

    # lowercase
    for col in rules.get("lowercase_fields", []):
        if col in df.columns:
            df[col] = df[col].astype(str).str.lower().replace({"none": None})

    # titlecase
    for col in rules.get("titlecase_fields", []):
        if col in df.columns:
            df[col] = df[col].astype(str).str.title().replace({"None": None})

    # parse dates
    for col in rules.get("date_fields", []):
        if col in df.columns:
            df[col] = df[col].apply(parse_date)

    return df

def validate_row(row: dict, validations: dict):
    """Validate a single row based on rules. Return (True, []) or (False, [errors])"""
    errors = []
    for field, rule in validations.items():
        val = row.get(field)
        # required
        if rule.get("required") and (val is None or (isinstance(val, str) and val.strip() == "")):
            errors.append(f"{field} is required")
            continue
        # regex
        if "regex" in rule and val is not None:
            if not re.match(rule["regex"], str(val)):
                errors.append(f"{field} fails regex {rule['regex']}")
        # type check (e.g., number)
        if rule.get("type") == "number" and val is not None:
            try:
                float(val)
            except Exception:
                errors.append(f"{field} is not numeric")
    return (len(errors) == 0), errors

# -------------------------
# Mapping & Filling
# -------------------------
def apply_mapping(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    """Return DataFrame with columns renamed to target names using mapping dict:
       mapping = { "source_col_name": "target_col_name", ... }
    """
    df = df.copy()
    # Only keep source columns that exist
    available = {s: t for s, t in mapping.items() if s in df.columns}
    df = df[list(available.keys())]
    df = df.rename(columns=available)
    logging.info(f"Applied mapping. Columns now: {list(df.columns)}")
    return df

def fill_template(template_path: str, output_path: str, df: pd.DataFrame, start_row: int = 2, sheet_name: str = None, formatting: dict = None, dry_run: bool = False):
    """Fill an Excel template workbook with the rows in df.
       - start_row: row number in template to start writing data (1-based)
       - formatting: dict, e.g., {"Date": "DATE", "Amount": "CURRENCY"}
       - If dry_run=True, do not save the output workbook.
    """
    wb = load_workbook(template_path)
    ws = wb.active if sheet_name is None else wb[sheet_name]
    headers = [cell.value for cell in ws[start_row - 1]]  # header row values
    # If header row is empty, fallback to df.columns
    if not any(headers):
        headers = list(df.columns)

    # Ensure all required headers exist; if not, create at the end
    for col in df.columns:
        if col not in headers:
            headers.append(col)

    # Write header row (optional): ensure header row contains headers
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=start_row - 1, column=idx, value=h)

    # Fill rows
    write_row = start_row
    for _, row in df.iterrows():
        for col_idx, header in enumerate(headers, start=1):
            val = row.get(header)
            cell = ws.cell(row=write_row, column=col_idx, value=None)
            if pd.isna(val):
                cell.value = None
            else:
                # Apply formatting rules (basic)
                if formatting and header in formatting:
                    f = formatting[header]
                    if f == "DATE":
                        dt = parse_date(val)
                        if dt:
                            cell.value = dt
                            cell.number_format = numbers.FORMAT_DATE_XLSX15
                        else:
                            cell.value = str(val)
                    elif f == "CURRENCY":
                        try:
                            cell.value = float(val)
                            cell.number_format = '"$"#,##0.00'
                        except Exception:
                            cell.value = val
                    else:
                        cell.value = val
                else:
                    cell.value = val
        write_row += 1

    if dry_run:
        logging.info("Dry-run enabled: not saving the filled workbook.")
        return wb

    # Save output
    wb.save(output_path)
    logging.info(f"Saved filled workbook to {output_path}")
    return output_path

# -------------------------
# CLI & Main
# -------------------------
def build_argparser():
    p = argparse.ArgumentParser(description="Spreadsheet Filler - populate an Excel template from CSV/Excel/Google Sheets")
    p.add_argument("--source", required=True, help="Path to source file (.csv or .xlsx) OR 'gsheet' for Google Sheets")
    p.add_argument("--template", required=True, help="Path to Excel template (.xlsx) to fill")
    p.add_argument("--out", required=True, help="Output Excel filepath (e.g., filled_output.xlsx)")
    p.add_argument("--mapping", required=True, help="Path to JSON mapping file (source_col -> target_col)")
    p.add_argument("--config", default=None, help="Optional JSON config with cleaning/validation rules")
    p.add_argument("--start-row", type=int, default=2, help="Row number to start writing data in template (1-based)")
    p.add_argument("--dry-run", action="store_true", help="If set, do not save output; just validate and preview")
    # Google Sheets extras:
    p.add_argument("--gsheet-key", help="If using Google Sheets, the spreadsheet key")
    p.add_argument("--gsheet-sheet", default="Sheet1", help="Worksheet name in Google Sheets")
    p.add_argument("--gsheet-creds", help="Path to Google service account JSON credentials (if using gsheet)")
    return p

def main():
    parser = build_argparser()
    args = parser.parse_args()

    # Load mapping
    mapping = read_config(args.mapping)

    # Optional config
    cfg = {}
    if args.config:
        cfg = read_config(args.config)

    # Load source
    if args.source.lower() == "gsheet":
        if not GS_SUPPORTED:
            logging.error("Google Sheets support not available. Install gspread and oauth2client.")
            sys.exit(1)
        if not args.gsheet_key or not args.gsheet_creds:
            logging.error("gsheet mode requires --gsheet-key and --gsheet-creds")
            sys.exit(1)
        df = load_google_sheet(args.gsheet_key, args.gsheet_sheet, args.gsheet_creds)
    else:
        df = load_source_file(args.source)

    # Clean data
    df = run_cleaning(df, cfg.get("cleaning", {}))

    # Apply mapping
    df_mapped = apply_mapping(df, mapping)

    # Validate rows
    validations = cfg.get("validations", {})
    bad_rows = []
    for i, row in df_mapped.iterrows():
        ok, errors = validate_row(row.to_dict(), validations)
        if not ok:
            bad_rows.append((int(i) + 1, errors))
            logging.warning(f"Row {i+1} validation errors: {errors}")

    if bad_rows:
        logging.error(f"Found {len(bad_rows)} rows with validation errors. Fix source or change validation rules.")
        if args.dry_run:
            logging.info("Dry-run: continuing despite validation errors.")
        else:
            sys.exit(1)

    # Backup template & create output folder
    if os.path.exists(args.out):
        backup_file(args.out)

    out_dir = Path(args.out).parent
    out_dir.mkdir(parents=True, exist_ok=True)

    # Fill template
    formatting = cfg.get("formatting", {})
    result = fill_template(
        template_path=args.template,
        output_path=args.out,
        df=df_mapped,
        start_row=args.start_row,
        formatting=formatting,
        dry_run=args.dry_run
    )

    logging.info("Spreadsheet filler completed.")

if __name__ == "__main__":
    main()
